#!/usr/bin/env python3
"""Compare EPA-certified 1800 RPM nonroad CI engines with the live catalog.

Usage:
  python3 data/analyze-epa-1800rpm.py \
    --epa-xlsx "/path/to/nonroad-compression-ignition-2011-present.xlsx" \
    --engines-json /tmp/haifeng-engines.json \
    --output-dir reports/epa-certification
"""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from collections import defaultdict
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl


MANUFACTURER_BRANDS = {
    "AB Volvo Penta": {"Volvo Penta"},
    "Caterpillar Inc.": {"Caterpillar"},
    "Cummins Inc.": {"Cummins"},
    "Daedong Corporation": {"Daedong"},
    "Deere & Company": {"John Deere"},
    "Detroit Diesel Corporation": {"Detroit Diesel"},
    "Deutz AG": {"Deutz"},
    "Discovery Energy, LLC.": {"Kohler"},
    "FAW JIEFANG AUTOMOTIVE CO.,LTD,WUXI DIESEL ENGINE WORKS": {"FAWDE"},
    "FPT Industrial S.p.A.": {"FPT"},
    "HD Construction Equipment Co., Ltd.": {"Hyundai"},
    "IHI Agri-Tech Corporation": {"Perkins", "Shibaura"},
    "Isuzu Motors Limited": {"Isuzu"},
    "Kirloskar Americas Corporation": {"Kirloskar"},
    "Komatsu Ltd.": {"Komatsu"},
    "Kubota Corporation": {"Kubota"},
    "Liebherr Machines Bulle SA": {"Kohler", "Liebherr"},
    "Lister Petter Limited": {"Lister Petter"},
    "MAN Truck & Bus AG": {"MAN"},
    "Mitsubishi Heavy Industries Engine & Turbocharger, Ltd.": {"Mitsubishi"},
    "Motorenfabrik Hatz GmbH & Co. KG": {"Hatz"},
    "Perkins Engines Co Ltd": {"Perkins"},
    "Rolls-Royce Solutions America Inc": {"MTU"},
    "Scania CV AB": {"Scania"},
    "Societe Internationale des Moteurs-Baudouin": {"Baudouin"},
    "Tianjin Lovol Engines Co., Ltd.": {"Lovol"},
    "Weichai Power Co.,Ltd.": {"Weichai"},
    "Yanmar Power Technology Co., Ltd.": {"Yanmar"},
    "Zhejiang Xinchai Co., Ltd.": {"Xinchai"},
}

FAMILY_MATCH_BRANDS = {
    "Caterpillar Inc.": {"Perkins"},
    "Cummins Inc.": {"Cummins"},
    "Perkins Engines Co Ltd": {"Perkins"},
    "Yanmar Power Technology Co., Ltd.": {"Yanmar"},
}

# Reviewed short certification families that intentionally represent multiple
# commercial generator-drive variants. The emissions token prevents an
# unregulated model with the same family prefix from satisfying EPA coverage.
REVIEWED_FAMILY_MATCHES = {
    ("Cummins Inc.", "QSL"): ("Cummins", "QSL9", "U.S. EPA Tier 3"),
    ("Deere & Company", "3029"): ("John Deere", "3029", "U.S. EPA"),
    ("Deere & Company", "4045"): ("John Deere", "4045", "U.S. EPA"),
    ("Deere & Company", "6068"): ("John Deere", "6068", "U.S. EPA"),
    ("Deere & Company", "6090"): ("John Deere", "6090", "U.S. EPA"),
    ("Deere & Company", "6136"): ("John Deere", "6136", "U.S. EPA"),
    (
        "Societe Internationale des Moteurs-Baudouin",
        "6M33",
    ): ("Baudouin", "6M33", "U.S. EPA Tier 2"),
    (
        "Societe Internationale des Moteurs-Baudouin",
        "8M33",
    ): ("Baudouin", "8M33", "U.S. EPA Tier 2"),
    (
        "Societe Internationale des Moteurs-Baudouin",
        "12M33",
    ): ("Baudouin", "12M33", "U.S. EPA Tier 2"),
    (
        "Societe Internationale des Moteurs-Baudouin",
        "12M55",
    ): ("Baudouin", "12M55", "U.S. EPA Tier 2"),
    (
        "Societe Internationale des Moteurs-Baudouin",
        "16M33",
    ): ("Baudouin", "16M33", "U.S. EPA Tier 2"),
    (
        "Societe Internationale des Moteurs-Baudouin",
        "16M55",
    ): ("Baudouin", "16M55", "U.S. EPA Tier 2"),
    (
        "Societe Internationale des Moteurs-Baudouin",
        "20M33",
    ): ("Baudouin", "20M33", "U.S. EPA Tier 2"),
    (
        "Perkins Engines Co Ltd",
        "C2P2",
    ): ("Perkins", "404J", "U.S. EPA Final Tier 4"),
    (
        "Perkins Engines Co Ltd",
        "C7P1",
    ): ("Perkins", "1206F", "U.S. EPA Final Tier 4"),
    (
        "Weichai Power Co.,Ltd.",
        "6M33",
    ): ("Baudouin", "6M33", "U.S. EPA Tier 2"),
    (
        "Weichai Power Co.,Ltd.",
        "12M33",
    ): ("Baudouin", "12M33", "U.S. EPA Tier 2"),
    (
        "Weichai Power Co.,Ltd.",
        "16M33",
    ): ("Baudouin", "16M33", "U.S. EPA Tier 2"),
}

CERTIFICATION_ALIASES = {
    ("IHI Agri-Tech Corporation", "402F05"): (
        "Perkins",
        "402F-05(C0.5)",
    ),
    ("IHI Agri-Tech Corporation", "403D11"): ("Shibaura", "S773L-F"),
    ("IHI Agri-Tech Corporation", "403D15"): ("Shibaura", "N843-F"),
    ("IHI Agri-Tech Corporation", "403F07"): ("Shibaura", "E673L-F"),
    ("IHI Agri-Tech Corporation", "403F11"): ("Shibaura", "S773L-F"),
    ("IHI Agri-Tech Corporation", "403F15"): ("Shibaura", "N843-F"),
    ("IHI Agri-Tech Corporation", "404D22"): ("Perkins", "404D-22G"),
    ("IHI Agri-Tech Corporation", "404D22TA"): ("Perkins", "404D-22TAG"),
    ("IHI Agri-Tech Corporation", "404D22TAC2P2"): (
        "Perkins",
        "404D-22TAG",
    ),
    ("IHI Agri-Tech Corporation", "C0P5"): (
        "Perkins",
        "402F-05(C0.5)",
    ),
    ("IHI Agri-Tech Corporation", "C0P7"): ("Shibaura", "E673L-F"),
    ("IHI Agri-Tech Corporation", "C1P1"): ("Shibaura", "S773L-F"),
    ("IHI Agri-Tech Corporation", "C1P5"): ("Shibaura", "N843-F"),
    ("IHI Agri-Tech Corporation", "361800C"): ("Shibaura", "N4LDI-TA"),
    ("IHI Agri-Tech Corporation", "4LT301800C"): ("Shibaura", "N4LDI-T"),
    ("IHI Agri-Tech Corporation", "ER49DI1800C"): (
        "Perkins",
        "404F-E22TA",
    ),
    ("IHI Agri-Tech Corporation", "ER49DI1800CCD"): (
        "Perkins",
        "404F-E22TA",
    ),
    ("IHI Agri-Tech Corporation", "101800C"): ("Shibaura", "E673L-C"),
    ("IHI Agri-Tech Corporation", "271800C"): ("Shibaura", "N844-D"),
    ("IHI Agri-Tech Corporation", "3LT201800C"): ("Shibaura", "N3LDI-T"),
    ("IHI Agri-Tech Corporation", "GG71800C"): (
        "Perkins",
        "402F-05(C0.5)",
    ),
    ("IHI Agri-Tech Corporation", "GH101800C"): ("Shibaura", "E673L-C"),
    ("IHI Agri-Tech Corporation", "GK221800C"): ("Shibaura", "N843-F"),
    ("IHI Agri-Tech Corporation", "151800C"): ("Shibaura", "S773L-D"),
    ("IHI Agri-Tech Corporation", "2003"): ("Shibaura", "N843-D"),
    ("IHI Agri-Tech Corporation", "211800C"): ("Shibaura", "N843-D"),
    ("IHI Agri-Tech Corporation", "321800C"): ("Shibaura", "N844L-D"),
    ("IHI Agri-Tech Corporation", "E3618C"): ("Shibaura", "E673L-F"),
    ("IHI Agri-Tech Corporation", "E3718C"): ("Shibaura", "E673L-F"),
    ("IHI Agri-Tech Corporation", "EG51800C"): (
        "Perkins",
        "402F-05(C0.5)",
    ),
    ("IHI Agri-Tech Corporation", "EH81800C"): ("Shibaura", "E673L-F"),
    ("IHI Agri-Tech Corporation", "EH8P1800C"): ("Shibaura", "E673L-F"),
    ("IHI Agri-Tech Corporation", "GJ131500CA"): ("Shibaura", "S773L-F"),
    ("IHI Agri-Tech Corporation", "GJ131800C"): ("Shibaura", "S773L-F"),
    ("IHI Agri-Tech Corporation", "GJ161800C"): ("Shibaura", "S773L-F"),
    ("IHI Agri-Tech Corporation", "GN291800C"): ("Perkins", "404D-22G"),
    ("IHI Agri-Tech Corporation", "GN331800C"): ("Perkins", "404D-22G"),
    ("IHI Agri-Tech Corporation", "GR491800C"): ("Perkins", "404D-22TAG"),
    ("IHI Agri-Tech Corporation", "EG41800C"): (
        "Perkins",
        "402F-05(C0.5)",
    ),
    ("IHI Agri-Tech Corporation", "EG61800C"): ("Shibaura", "E673L-F"),
    ("IHI Agri-Tech Corporation", "EH61800C"): ("Shibaura", "E673L-F"),
    (
        "Discovery Energy, LLC.",
        "KSDNATG140318",
    ): ("Kohler", "KSD1403NA"),
    ("FPT Industrial S.p.A.", "F2CCA615AH"): ("FPT", "C87 TE3F"),
    ("FPT Industrial S.p.A.", "F2CCP615AH"): ("FPT", "C87 TE1PV"),
    ("FPT Industrial S.p.A.", "F2CE9685AE"): ("FPT", "C87 TE3F"),
    ("FPT Industrial S.p.A.", "F2CE9685CE"): ("FPT", "C87 TE3F"),
    ("FPT Industrial S.p.A.", "F4GE9455AJ"): ("FPT", "NEF45SM2X"),
    ("FPT Industrial S.p.A.", "F4GE9455BJ"): ("FPT", "NEF45SM1X"),
    ("FPT Industrial S.p.A.", "F4GE9685AJ"): ("FPT", "NEF67 TM1X"),
    ("FPT Industrial S.p.A.", "F4HE0485BJ"): ("FPT", "NEF45 TE2P"),
    ("FPT Industrial S.p.A.", "F4HE0485CJ"): ("FPT", "NEF45 TE1P"),
    ("FPT Industrial S.p.A.", "F4HE0685AJ"): ("FPT", "NEF67 TE3PV"),
    ("FPT Industrial S.p.A.", "F4HE0685FJ"): ("FPT", "NEF67 TE2PV"),
    ("FPT Industrial S.p.A.", "F4HE0685GJ"): ("FPT", "NEF67 TE1PV"),
    ("Cummins Inc.", "QSK78G"): ("Cummins", "QSK78"),
    ("Cummins Inc.", "QST30G"): ("Cummins", "QST30"),
    ("Cummins Inc.", "QSX15G"): ("Cummins", "QSX15"),
    ("Cummins Inc.", "QSB5G7"): ("Cummins", "QSB4.5"),
    ("Cummins Inc.", "QSB5G8"): ("Cummins", "QSB4.5"),
    ("Cummins Inc.", "QSB5G9"): ("Cummins", "QSB4.5"),
    ("Deere & Company", "6135H"): ("John Deere", "6135"),
    ("AB Volvo Penta", "TAD1641GE"): ("Volvo Penta", "TAD1641GE-B"),
    ("AB Volvo Penta", "TAD1642GE"): ("Volvo Penta", "TAD1642GE-B"),
    ("Discovery Energy, LLC.", "KD2504ESM"): ("Kohler", "KDI2504ESM"),
    ("Discovery Energy, LLC.", "LDW1603GE1800RPM"): (
        "Kohler",
        "KDW1603GE (1800 RPM)",
    ),
    ("Discovery Energy, LLC.", "LDW2204TGE"): ("Kohler", "KDW2204TGE"),
    ("Detroit Diesel Corporation", "SERIES60"): (
        "Detroit Diesel",
        "Series 60 14.0L",
    ),
    ("Detroit Diesel Corporation", "SERIES6014L"): (
        "Detroit Diesel",
        "Series 60 14.0L",
    ),
    (
        "Mitsubishi Heavy Industries Engine & Turbocharger, Ltd.",
        "S12A2PTAW",
    ): ("Mitsubishi", "S12A2-Y2PTAW-2"),
    (
        "Mitsubishi Heavy Industries Engine & Turbocharger, Ltd.",
        "S12HPTAW",
    ): ("Mitsubishi", "S12H-Y2PTAW-1"),
    (
        "Mitsubishi Heavy Industries Engine & Turbocharger, Ltd.",
        "S12RPTAW",
    ): ("Mitsubishi", "S12R-Y2PTAW-1"),
    (
        "Mitsubishi Heavy Industries Engine & Turbocharger, Ltd.",
        "S16RPTAW",
    ): ("Mitsubishi", "S16R-Y2PTAW-1"),
    (
        "Mitsubishi Heavy Industries Engine & Turbocharger, Ltd.",
        "S16RPTAW2",
    ): ("Mitsubishi", "S16R-Y2PTAW2"),
    (
        "Mitsubishi Heavy Industries Engine & Turbocharger, Ltd.",
        "S6RPTAW",
    ): ("Mitsubishi", "S6R-Y2PTAW"),
    ("Rolls-Royce Solutions America Inc", "10V1600G70S3B"): ("MTU", "10V1600G70S"),
    ("Rolls-Royce Solutions America Inc", "10V1600G70S3D"): ("MTU", "10V1600G70S"),
    ("Rolls-Royce Solutions America Inc", "6R1600G10S"): (
        "MTU",
        "6R1600G10S 3B",
    ),
    ("Rolls-Royce Solutions America Inc", "6R1600G20S"): (
        "MTU",
        "6R1600G20S 3B",
    ),
    ("Rolls-Royce Solutions America Inc", "6R1600G70S"): (
        "MTU",
        "6R1600G70S 3D",
    ),
    ("Rolls-Royce Solutions America Inc", "6R1600G80S"): (
        "MTU",
        "6R1600G80S 3D",
    ),
    ("Rolls-Royce Solutions America Inc", "8V1600G10S"): (
        "MTU",
        "8V1600G10S 3B",
    ),
    ("Rolls-Royce Solutions America Inc", "8V1600G20S"): (
        "MTU",
        "8V1600G20S 3B",
    ),
    ("Rolls-Royce Solutions America Inc", "8V1600G70S"): (
        "MTU",
        "8V1600G70S 3D",
    ),
    ("Rolls-Royce Solutions America Inc", "8V1600G80S"): (
        "MTU",
        "8V1600G80S 3D",
    ),
    ("Rolls-Royce Solutions America Inc", "10V1600G20S3B"): (
        "MTU",
        "10V1600G20S",
    ),
    ("Rolls-Royce Solutions America Inc", "10V1600G80S3D"): (
        "MTU",
        "10V1600G80S",
    ),
    ("Rolls-Royce Solutions America Inc", "12V1600G10S3B"): (
        "MTU",
        "12V1600G10S",
    ),
    ("Rolls-Royce Solutions America Inc", "12V1600G20S3B"): (
        "MTU",
        "12V1600G20S",
    ),
    ("Rolls-Royce Solutions America Inc", "12V1600G70S3D"): (
        "MTU",
        "12V1600G70S",
    ),
    ("Rolls-Royce Solutions America Inc", "12V1600G80S3D"): (
        "MTU",
        "12V1600G80S",
    ),
    ("Rolls-Royce Solutions America Inc", "12V2000G443B"): ("MTU", "MTU 12V2000 G44"),
    ("Rolls-Royce Solutions America Inc", "12V2000G443D"): ("MTU", "MTU 12V2000 G44"),
    ("Rolls-Royce Solutions America Inc", "16V2000G443B"): ("MTU", "MTU 16V2000 G44"),
    ("Rolls-Royce Solutions America Inc", "16V2000G443D"): ("MTU", "MTU 16V2000 G44"),
    ("Rolls-Royce Solutions America Inc", "12V2000G45TB3B"): ("MTU", "MTU 12V2000 G45"),
    ("Rolls-Royce Solutions America Inc", "12V2000G45TB3D"): ("MTU", "MTU 12V2000 G45"),
    ("Rolls-Royce Solutions America Inc", "12V2000G45TD3B"): ("MTU", "MTU 12V2000 G45"),
    ("Rolls-Royce Solutions America Inc", "12V2000G45TD3D"): ("MTU", "MTU 12V2000 G45"),
    ("Rolls-Royce Solutions America Inc", "16V2000G45TB3B"): ("MTU", "MTU 16V2000 G45"),
    ("Rolls-Royce Solutions America Inc", "16V2000G45TB3D"): ("MTU", "MTU 16V2000 G45"),
    ("Rolls-Royce Solutions America Inc", "16V2000G45TD3B"): ("MTU", "MTU 16V2000 G45"),
    ("Rolls-Royce Solutions America Inc", "16V2000G45TD3D"): ("MTU", "MTU 16V2000 G45"),
    ("Rolls-Royce Solutions America Inc", "12V2000G843B"): ("MTU", "MTU 12V2000 G84"),
    ("Rolls-Royce Solutions America Inc", "12V2000G843D"): ("MTU", "MTU 12V2000 G84"),
    ("Rolls-Royce Solutions America Inc", "16V2000G843B"): ("MTU", "MTU 16V2000 G84"),
    ("Rolls-Royce Solutions America Inc", "16V2000G843D"): ("MTU", "MTU 16V2000 G84"),
    ("Rolls-Royce Solutions America Inc", "12V2000G85TB3B"): ("MTU", "MTU 12V2000 G85"),
    ("Rolls-Royce Solutions America Inc", "12V2000G85TB3D"): ("MTU", "MTU 12V2000 G85"),
    ("Rolls-Royce Solutions America Inc", "12V2000G85TD3B"): ("MTU", "MTU 12V2000 G85"),
    ("Rolls-Royce Solutions America Inc", "12V2000G85TD3D"): ("MTU", "MTU 12V2000 G85"),
    ("Rolls-Royce Solutions America Inc", "16V2000G85TB3B"): ("MTU", "MTU 16V2000 G85"),
    ("Rolls-Royce Solutions America Inc", "16V2000G85TB3D"): ("MTU", "MTU 16V2000 G85"),
    ("Rolls-Royce Solutions America Inc", "16V2000G85TD3B"): ("MTU", "MTU 16V2000 G85"),
    ("Rolls-Royce Solutions America Inc", "16V2000G85TD3D"): ("MTU", "MTU 16V2000 G85"),
    ("Rolls-Royce Solutions America Inc", "18V2000G85TB3B"): ("MTU", "MTU 18V2000 G85"),
    ("Rolls-Royce Solutions America Inc", "18V2000G85TB3D"): ("MTU", "MTU 18V2000 G85"),
    ("Rolls-Royce Solutions America Inc", "18V2000G85TD3B"): ("MTU", "MTU 18V2000 G85"),
    ("Rolls-Royce Solutions America Inc", "18V2000G85TD3D"): ("MTU", "MTU 18V2000 G85"),
    ("Rolls-Royce Solutions America Inc", "12V2000P823A"): ("MTU", "MTU 12V2000 P82"),
    ("Rolls-Royce Solutions America Inc", "12V2000P823B3C"): ("MTU", "MTU 12V2000 P82"),
    ("Rolls-Royce Solutions America Inc", "16V2000P823A"): ("MTU", "MTU 16V2000 P82"),
    ("Rolls-Royce Solutions America Inc", "16V2000P823B3C"): ("MTU", "MTU 16V2000 P82"),
    ("Rolls-Royce Solutions America Inc", "12V4000G433B"): ("MTU", "MTU 12V4000 G43"),
    ("Rolls-Royce Solutions America Inc", "12V4000G433D"): ("MTU", "MTU 12V4000 G43"),
    ("Rolls-Royce Solutions America Inc", "12V4000G433F"): ("MTU", "MTU 12V4000 G43"),
    ("Rolls-Royce Solutions America Inc", "12V4000G833B"): ("MTU", "MTU 12V4000 G83"),
    ("Rolls-Royce Solutions America Inc", "12V4000G833D"): ("MTU", "MTU 12V4000 G83"),
    ("Rolls-Royce Solutions America Inc", "12V4000G833F"): ("MTU", "MTU 12V4000 G83"),
    ("Rolls-Royce Solutions America Inc", "16V4000G433B"): ("MTU", "MTU 16V4000 G43"),
    ("Rolls-Royce Solutions America Inc", "16V4000G433D"): ("MTU", "MTU 16V4000 G43"),
    ("Rolls-Royce Solutions America Inc", "16V4000G433F"): ("MTU", "MTU 16V4000 G43"),
    ("Rolls-Royce Solutions America Inc", "16V4000G833B"): ("MTU", "MTU 16V4000 G83"),
    ("Rolls-Royce Solutions America Inc", "16V4000G833D"): ("MTU", "MTU 16V4000 G83"),
    ("Rolls-Royce Solutions America Inc", "16V4000G833F"): ("MTU", "MTU 16V4000 G83"),
    ("Rolls-Royce Solutions America Inc", "16V4000G83L3D"): ("MTU", "MTU 16V4000 G83L"),
    ("Rolls-Royce Solutions America Inc", "20V4000G433B"): ("MTU", "MTU 20V4000 G43"),
    ("Rolls-Royce Solutions America Inc", "20V4000G433D"): ("MTU", "MTU 20V4000 G43"),
    ("Rolls-Royce Solutions America Inc", "20V4000G433F"): ("MTU", "MTU 20V4000 G43"),
    ("Rolls-Royce Solutions America Inc", "20V4000G833B"): ("MTU", "MTU 20V4000 G83"),
    ("Rolls-Royce Solutions America Inc", "20V4000G833D"): ("MTU", "MTU 20V4000 G83"),
    ("Rolls-Royce Solutions America Inc", "20V4000G833F"): ("MTU", "MTU 20V4000 G83"),
    ("Rolls-Royce Solutions America Inc", "20V4000G83L3B"): ("MTU", "MTU 20V4000 G83L"),
    ("Rolls-Royce Solutions America Inc", "20V4000G83L3D"): ("MTU", "MTU 20V4000 G83L"),
    ("Rolls-Royce Solutions America Inc", "20V4000G83L3F"): ("MTU", "MTU 20V4000 G83L"),
    ("Liebherr Machines Bulle SA", "D9812G"): ("Liebherr", "D9812"),
    ("Liebherr Machines Bulle SA", "D9816G"): ("Liebherr", "D9816"),
    ("Liebherr Machines Bulle SA", "D9820G"): ("Liebherr", "D9820"),
    ("Liebherr Machines Bulle SA", "D976A702"): ("Liebherr", "D976"),
    ("Isuzu Motors Limited", "BV4LE1"): ("Isuzu", "KV-4LE1"),
    ("Isuzu Motors Limited", "BV4LE1T"): ("Isuzu", "KV-4LE1T"),
    ("Isuzu Motors Limited", "BV4LE2"): ("Isuzu", "LV-4LE2"),
    ("Kubota Corporation", "D1005BGEF"): ("Kubota", "D1005-E4BG1-SAE-2"),
    ("Kubota Corporation", "D1005EF"): ("Kubota", "D1005-E4BG1-SAE-2"),
    ("Kubota Corporation", "D1105BGEF"): ("Kubota", "D1105-E4BG1-SAE-2X"),
    ("Kubota Corporation", "D1105EF"): ("Kubota", "D1105-E4BG1-SAE-2X"),
    ("Kubota Corporation", "D1305BGEF"): ("Kubota", "D1305-E4BG1-CHN-1"),
    ("Kubota Corporation", "D1503MBGEF"): ("Kubota", "D1503-M-E4-BG"),
    ("Kubota Corporation", "D1703MBGET"): ("Kubota", "D1703-M-E3-BG"),
    ("Kubota Corporation", "D1803CRTIBGEF"): ("Kubota", "D1803-CR-TI-E4-BG"),
    ("Kubota Corporation", "V1505BGEF"): ("Kubota", "V1505-E4BG1-SAE-2X"),
    ("Kubota Corporation", "V2203MBGET"): ("Kubota", "V2203-M-E3-BG"),
    ("Kubota Corporation", "V2403CRTIBGEF"): ("Kubota", "V2403-CR-TI-E4-BG"),
    ("Kubota Corporation", "V3300BGET"): ("Kubota", "V3300-E3-BG"),
    ("Kubota Corporation", "V3600TBGET"): ("Kubota", "V3600-T-E3-BG"),
    ("Kubota Corporation", "V3800DITBGET"): ("Kubota", "V3800DI-T-E3-BG"),
    ("Kubota Corporation", "Z482D2EF"): ("Kubota", "Z482-E4B-CHN-1"),
    ("Kubota Corporation", "V3300BGEF"): ("Kubota", "V3300-E3-BG"),
    ("Kubota Corporation", "D1005BGET"): ("Kubota", "D1005-E4BG1-SAE-2"),
    ("Kubota Corporation", "D1105BGET"): ("Kubota", "D1105-E4BG1-SAE-2X"),
    ("Kubota Corporation", "D1305BGET"): ("Kubota", "D1305-E4BG1-CHN-1"),
    ("Kubota Corporation", "D1503MBGET"): ("Kubota", "D1503-M-E4-BG"),
    ("Kubota Corporation", "V1505BGET"): ("Kubota", "V1505-E4BG1-SAE-2X"),
    ("Perkins Engines Co Ltd", "C1P1"): ("Perkins", "S773L-F"),
    ("Perkins Engines Co Ltd", "C1P5"): ("Perkins", "403F-15"),
    ("Perkins Engines Co Ltd", "C0P5"): ("Perkins", "402F-05(C0.5)"),
    ("Perkins Engines Co Ltd", "C0P7"): ("Perkins", "403F-07(C0.7)"),
    ("Perkins Engines Co Ltd", "E673LF"): ("Perkins", "403F-07(C0.7)"),
    ("Perkins Engines Co Ltd", "403D11C1P1"): ("Perkins", "403D-11G"),
    ("Perkins Engines Co Ltd", "403F11C1P1"): ("Perkins", "403F-11G"),
    ("Perkins Engines Co Ltd", "403F15C1P5"): ("Perkins", "403F-15"),
    ("Perkins Engines Co Ltd", "404FE22TAC2P2"): ("Perkins", "404F-E22TA"),
    ("Perkins Engines Co Ltd", "404D22TAC2P2"): ("Perkins", "404D-22TAG"),
    ("Perkins Engines Co Ltd", "404JE22TAC2P2"): ("Perkins", "404J-E22TAG"),
    ("Perkins Engines Co Ltd", "1104D44TC4P4"): ("Perkins", "1104D-44TG1"),
    ("Perkins Engines Co Ltd", "1104DE44TC4P4"): ("Perkins", "1104D-E44TG1"),
    ("Perkins Engines Co Ltd", "1104DE44TAC4P4"): ("Perkins", "1104D-E44TAG2"),
    ("Perkins Engines Co Ltd", "1106DE70TAC7P1"): ("Perkins", "1106D-E70TAG5"),
    ("Perkins Engines Co Ltd", "1204JE44TTAC4P4"): ("Perkins", "1204J-E44TTAG2"),
    ("Perkins Engines Co Ltd", "1206JE70TTAC7P1"): ("Perkins", "1206J-E70TTAG4"),
    ("Perkins Engines Co Ltd", "36621800"): ("Perkins", "1204E-E44TA(C4.4)"),
    ("Perkins Engines Co Ltd", "37901800"): ("Perkins", "1204E-E44TTA(C4.4)"),
    ("Perkins Engines Co Ltd", "37921800"): ("Perkins", "1204E-E44TA(C4.4)"),
    ("Perkins Engines Co Ltd", "38621800"): ("Perkins", "1204E-E44TTA(C4.4)"),
    ("Yanmar Power Technology Co., Ltd.", "3TNGAG"): ("Yanmar", "4TNV98C-GGE"),
    ("Yanmar Power Technology Co., Ltd.", "3TNGA"): ("Yanmar", "4TNV98C-GGE"),
    ("Yanmar Power Technology Co., Ltd.", "3TNGP"): ("Yanmar", "4TNV98C-GGE"),
    ("Yanmar Power Technology Co., Ltd.", "3TNV88CL"): ("Yanmar", "3TNV88F-UG6GE"),
    ("Yanmar Power Technology Co., Ltd.", "3CB1G"): ("Yanmar", "3TNV76-CL"),
    ("Yanmar Power Technology Co., Ltd.", "3JTGA"): ("Yanmar", "3JTGP1"),
    ("Yanmar Power Technology Co., Ltd.", "3JTGAK"): ("Yanmar", "3JTGP1"),
    ("Yanmar Power Technology Co., Ltd.", "3KNGA"): (
        "Yanmar",
        "3TNV88F-UG6GE",
    ),
    ("Yanmar Power Technology Co., Ltd.", "3MTGAK"): ("Yanmar", "3MTGAG"),
    ("Yanmar Power Technology Co., Ltd.", "3TTGAG"): ("Yanmar", "4TNV98CT-GGE"),
    ("Yanmar Power Technology Co., Ltd.", "3TTGA"): ("Yanmar", "4TNV98CT-GGE"),
    ("Yanmar Power Technology Co., Ltd.", "3TTGP"): ("Yanmar", "4TNV98CT-GGE"),
    ("Yanmar Power Technology Co., Ltd.", "4TNGAC"): ("Yanmar", "4TNV98C-GGE"),
    ("Yanmar Power Technology Co., Ltd.", "4TNGPC"): ("Yanmar", "4TNV98C-GGE"),
    ("Yanmar Power Technology Co., Ltd.", "4RTGAC"): ("Yanmar", "4TNV86CT"),
    ("Yanmar Power Technology Co., Ltd.", "4RTGPC"): ("Yanmar", "4TNV86CT"),
    ("Yanmar Power Technology Co., Ltd.", "4TTGAC"): ("Yanmar", "4TNV98CT-GGE"),
    ("Yanmar Power Technology Co., Ltd.", "4TTGPC"): ("Yanmar", "4TNV98CT-GGE"),
    ("Yanmar Power Technology Co., Ltd.", "3NNGAG"): ("Yanmar", "4TNV88-CL"),
    ("Yanmar Power Technology Co., Ltd.", "3NNGA"): ("Yanmar", "4TNV88-CL"),
    ("Yanmar Power Technology Co., Ltd.", "4HNGAM"): (
        "Yanmar",
        "3TNV88F-UG6GE",
    ),
    ("Motorenfabrik Hatz GmbH & Co. KG", "1D81SZTU"): ("Hatz", "1D81"),
    (
        "Yanmar Power Technology Co., Ltd.",
        "4HNGFM",
    ): ("Yanmar", "3TNV88F-UG6GE"),
    (
        "Yanmar Power Technology Co., Ltd.",
        "4HNGPM",
    ): ("Yanmar", "3TNV88F-UG6GE"),
    (
        "Yanmar Power Technology Co., Ltd.",
        "4WNGAA",
    ): ("Yanmar", "3TNV80F-NG6GE"),
    (
        "Yanmar Power Technology Co., Ltd.",
        "4WNGPA",
    ): ("Yanmar", "3TNV80F-NG6GE"),
    (
        "Yanmar Power Technology Co., Ltd.",
        "5ENGAA",
    ): ("Yanmar", "3TNM74F-NG6GE"),
}

# Some EPA model fields are certification-group identifiers rather than engine
# model names. A group is represented only when every reviewed commercial model
# listed here exists in the catalog.
CERTIFICATION_ALIAS_GROUPS = {
    ("IHI Agri-Tech Corporation", "C2P2"): (
        ("Perkins", "404D-22G"),
        ("Perkins", "404D-22TAG"),
        ("Perkins", "404F-E22TA"),
    ),
    ("Lister Petter Limited", "408"): (
        ("Lister Petter", "LPWS2"),
        ("Lister Petter", "LPWS3"),
        ("Lister Petter", "LPWS4"),
        ("Lister Petter", "LPWST4"),
    ),
    ("Lister Petter Limited", "418"): (
        ("Lister Petter", "LPWS2"),
    ),
    ("Lister Petter Limited", "443"): (
        ("Lister Petter", "LPWS2"),
        ("Lister Petter", "LPWS3"),
        ("Lister Petter", "LPWS4"),
        ("Lister Petter", "LPWST4"),
    ),
    ("Lister Petter Limited", "458"): (
        ("Lister Petter", "LPWS2"),
        ("Lister Petter", "LPWS3"),
        ("Lister Petter", "LPWS4"),
        ("Lister Petter", "LPWST4"),
    ),
    ("Lister Petter Limited", "468"): (
        ("Lister Petter", "LPWS2"),
        ("Lister Petter", "LPWS3"),
        ("Lister Petter", "LPWS4"),
    ),
    ("Lister Petter Limited", "474"): (
        ("Lister Petter", "LPWS2"),
        ("Lister Petter", "LPWS3"),
        ("Lister Petter", "LPWS4"),
        ("Lister Petter", "LPWST4"),
    ),
    ("Lister Petter Limited", "479"): (
        ("Lister Petter", "LPWS2"),
        ("Lister Petter", "LPWS3"),
        ("Lister Petter", "LPWS4"),
        ("Lister Petter", "LPWST4"),
    ),
}

REPRESENTED_STATUSES = {
    "exact_brand_match",
    "brand_prefix_match",
    "base_brand_match",
    "certification_trim_match",
    "certification_alias_match",
    "certification_alias_group_match",
    "family_brand_match",
}


def normalize_model(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).upper()
    # Decimal displacement designations are significant: C1.5 is not C15.
    text = re.sub(r"(?<=\d)\.(?=\d)", "P", text)
    return re.sub(r"[^A-Z0-9]+", "", text)


def sorted_values(values: set[object]) -> list[object]:
    return sorted(value for value in values if value not in (None, ""))


def read_family_info(workbook: openpyxl.Workbook) -> dict[tuple[int, str], dict]:
    sheet = workbook["Family Info"]
    headers = [cell.value for cell in sheet[2]]
    index = {header: position for position, header in enumerate(headers)}
    families = {}

    for row in sheet.iter_rows(min_row=3, values_only=True):
        year = row[index["Model Year"]]
        family = row[index["Engine Family"]]
        if year is None or family is None:
            continue
        families[(int(year), str(family))] = {
            "manufacturer": row[index["Manufacturer"]],
            "certificate": row[index["Certificate #"]],
            "applicable_tier": row[index["Applicable Tier"]],
            "compliance_standard": row[index["Applicable Compliance Standard"]],
            "fuel": row[index["Fuel"]],
            "engine_operation": row[index["Engine Operation"]],
        }
    return families


def read_1800_rpm_models(
    workbook: openpyxl.Workbook, families: dict[tuple[int, str], dict]
) -> tuple[dict[tuple[str, str], dict], int]:
    sheet = workbook["Model Info"]
    headers = [cell.value for cell in sheet[2]]
    index = {header: position for position, header in enumerate(headers)}
    models: dict[tuple[str, str], dict] = {}
    source_rows = 0

    for row in sheet.iter_rows(min_row=3, values_only=True):
        rated_speed = row[index["Rated Speed (RPM)"]]
        if not isinstance(rated_speed, (int, float)) or abs(rated_speed - 1800) > 0.01:
            continue

        source_rows += 1
        year = int(row[index["Model Year"]])
        family = str(row[index["Engine Family"]])
        family_info = families[(year, family)]
        manufacturer = str(family_info["manufacturer"])
        display_model = str(row[index["Engine Model"]]).strip()
        normalized_model = normalize_model(display_model)
        key = (manufacturer, normalized_model)

        if key not in models:
            models[key] = {
                "manufacturer": manufacturer,
                "epa_model": display_model,
                "normalized_model": normalized_model,
                "model_years": set(),
                "engine_families": set(),
                "engine_codes": set(),
                "certificates": set(),
                "applicable_tiers": set(),
                "compliance_standards": set(),
                "fuels": set(),
                "engine_operations": set(),
                "rated_powers_kw": set(),
                "displacements_l": set(),
            }

        record = models[key]
        record["model_years"].add(year)
        record["engine_families"].add(family)
        record["engine_codes"].add(row[index["Engine Code"]])
        record["certificates"].add(family_info["certificate"])
        record["applicable_tiers"].add(family_info["applicable_tier"])
        record["compliance_standards"].add(family_info["compliance_standard"])
        record["fuels"].add(family_info["fuel"])
        record["engine_operations"].add(family_info["engine_operation"])
        record["rated_powers_kw"].add(row[index["Rated Power (KW)"]])
        record["displacements_l"].add(row[index["Total Displacement"]])

    return models, source_rows


def build_database_indexes(engines: list[dict]) -> tuple[dict, dict, dict]:
    by_model: dict[str, list[dict]] = defaultdict(list)
    by_model_without_brand_prefix: dict[str, list[dict]] = defaultdict(list)
    by_brand: dict[str, list[dict]] = defaultdict(list)
    for engine in engines:
        normalized = normalize_model(engine["model"])
        engine = {**engine, "normalized_model": normalized}
        by_model[normalized].append(engine)
        by_brand[engine["brand"]].append(engine)
        normalized_brand = normalize_model(engine["brand"])
        if (
            normalized.startswith(normalized_brand)
            and len(normalized) - len(normalized_brand) >= 5
        ):
            without_prefix = normalized[len(normalized_brand):]
            by_model_without_brand_prefix[without_prefix].append(engine)
    return by_model, by_model_without_brand_prefix, by_brand


def probable_brand_matches(
    normalized_model: str, brands: set[str], by_brand: dict[str, list[dict]]
) -> list[dict]:
    candidates = []
    for brand in brands:
        for engine in by_brand.get(brand, []):
            database_model = engine["normalized_model"]
            if min(len(normalized_model), len(database_model)) < 5:
                continue
            ratio = SequenceMatcher(None, normalized_model, database_model).ratio()
            prefix_match = (
                normalized_model.startswith(database_model)
                or database_model.startswith(normalized_model)
            ) and abs(len(normalized_model) - len(database_model)) <= 4
            if ratio >= 0.88 or prefix_match:
                candidates.append(
                    {
                        "brand": engine["brand"],
                        "model": engine["model"],
                        "slug": engine["slug"],
                        "similarity": round(ratio, 3),
                    }
                )
    return sorted(
        candidates,
        key=lambda candidate: (-candidate["similarity"], candidate["slug"]),
    )[:5]


def certification_trim_candidates(
    manufacturer: str,
    normalized_model: str,
    by_brand: dict[str, list[dict]],
) -> list[dict]:
    candidates = []
    for brand, suffix_pattern in (
        ("Kohler", r"A?6(?:[A-D]ES|[A-C]N[CPS])"),
        ("Liebherr", r"A70[24]"),
    ):
        if manufacturer != "Liebherr Machines Bulle SA":
            continue
        for engine in by_brand.get(brand, []):
            base = engine["normalized_model"]
            if (
                len(base) >= 5
                and normalized_model.startswith(base)
                and re.fullmatch(suffix_pattern, normalized_model[len(base):])
            ):
                candidates.append(engine)
    return sorted(candidates, key=lambda engine: engine["slug"])


def match_models(models: dict, engines: list[dict]) -> list[dict]:
    by_model, by_model_without_brand_prefix, by_brand = build_database_indexes(engines)
    results = []

    for record in models.values():
        manufacturer = record["manufacturer"]
        normalized_model = record["normalized_model"]
        expected_brands = MANUFACTURER_BRANDS.get(manufacturer, set())
        exact_candidates = by_model.get(normalized_model, [])
        exact_brand_candidates = [
            engine
            for engine in exact_candidates
            if engine["brand"] in expected_brands
        ]
        brand_prefix_candidates = [
            engine
            for engine in by_model_without_brand_prefix.get(normalized_model, [])
            if engine["brand"] in expected_brands
        ]
        base_model = normalize_model(record["epa_model"].split("/", 1)[0])
        base_brand_candidates = [
            engine
            for engine in by_model.get(base_model, [])
            if engine["brand"] in expected_brands
        ] if "/" in record["epa_model"] and len(base_model) >= 5 else []
        trim_candidates = certification_trim_candidates(
            manufacturer,
            normalized_model,
            by_brand,
        )
        alias_target = CERTIFICATION_ALIASES.get(
            (manufacturer, normalized_model)
        )
        certification_alias_candidates = [
            engine
            for engine in by_model.get(
                normalize_model(alias_target[1]) if alias_target else "",
                [],
            )
            if alias_target and engine["brand"] == alias_target[0]
        ]
        alias_group = CERTIFICATION_ALIAS_GROUPS.get(
            (manufacturer, normalized_model)
        )
        certification_alias_group_candidates = []
        if alias_group:
            group_matches = []
            for group_brand, group_model in alias_group:
                target_matches = [
                    engine
                    for engine in by_model.get(normalize_model(group_model), [])
                    if engine["brand"] == group_brand
                ]
                if not target_matches:
                    group_matches = []
                    break
                group_matches.extend(target_matches)
            certification_alias_group_candidates = sorted(
                group_matches,
                key=lambda engine: engine["slug"],
            )
        reviewed_family = REVIEWED_FAMILY_MATCHES.get(
            (manufacturer, normalized_model)
        )
        if reviewed_family:
            family_brand, family_prefix, emissions_token = reviewed_family
            normalized_prefix = normalize_model(family_prefix)
            family_brand_candidates = sorted(
                [
                    engine
                    for engine in by_brand.get(family_brand, [])
                    if engine["normalized_model"].startswith(normalized_prefix)
                    and emissions_token in str(
                        engine.get("emissions_standard") or ""
                    )
                ],
                key=lambda engine: engine["slug"],
            )[:10]
        else:
            family_brands = FAMILY_MATCH_BRANDS.get(manufacturer, set())
            family_brand_candidates = sorted(
                [
                    engine
                    for brand in family_brands
                    for engine in by_brand.get(brand, [])
                    if len(normalized_model) >= 5
                    and engine["normalized_model"].startswith(normalized_model)
                    and len(engine["normalized_model"]) > len(normalized_model)
                    and len(engine["normalized_model"]) - len(normalized_model) <= 12
                ],
                key=lambda engine: engine["slug"],
            )[:10]

        if exact_brand_candidates:
            status = "exact_brand_match"
            matches = exact_brand_candidates
            probable = []
        elif brand_prefix_candidates:
            status = "brand_prefix_match"
            matches = brand_prefix_candidates
            probable = []
        elif base_brand_candidates:
            status = "base_brand_match"
            matches = base_brand_candidates
            probable = []
        elif trim_candidates:
            status = "certification_trim_match"
            matches = trim_candidates
            probable = []
        elif certification_alias_candidates:
            status = "certification_alias_match"
            matches = certification_alias_candidates
            probable = []
        elif certification_alias_group_candidates:
            status = "certification_alias_group_match"
            matches = certification_alias_group_candidates
            probable = []
        elif family_brand_candidates:
            status = "family_brand_match"
            matches = family_brand_candidates
            probable = []
        elif exact_candidates:
            status = "exact_model_other_brand"
            matches = exact_candidates
            probable = []
        else:
            status = "not_found"
            matches = []
            probable = probable_brand_matches(
                normalized_model, expected_brands, by_brand
            )

        years = sorted_values(record["model_years"])
        powers = sorted_values(record["rated_powers_kw"])
        displacements = sorted_values(record["displacements_l"])
        results.append(
            {
                "manufacturer": manufacturer,
                "mapped_database_brands": sorted(expected_brands),
                "epa_model": record["epa_model"],
                "normalized_model": normalized_model,
                "first_model_year": years[0],
                "latest_model_year": years[-1],
                "model_years": years,
                "engine_families": sorted_values(record["engine_families"]),
                "engine_codes": sorted_values(record["engine_codes"]),
                "certificates": sorted_values(record["certificates"]),
                "applicable_tiers": sorted_values(record["applicable_tiers"]),
                "compliance_standards": sorted_values(
                    record["compliance_standards"]
                ),
                "fuels": sorted_values(record["fuels"]),
                "engine_operations": sorted_values(record["engine_operations"]),
                "rated_power_kw_min": powers[0] if powers else None,
                "rated_power_kw_max": powers[-1] if powers else None,
                "displacements_l": displacements,
                "match_status": status,
                "database_matches": [
                    {
                        "brand": engine["brand"],
                        "model": engine["model"],
                        "slug": engine["slug"],
                        "rpm_rated": engine["rpm_rated"],
                        "emissions_standard": engine["emissions_standard"],
                        "fuel_type": engine["fuel_type"],
                    }
                    for engine in matches
                ],
                "probable_database_matches": probable,
            }
        )

    return sorted(
        results,
        key=lambda result: (
            result["match_status"] not in REPRESENTED_STATUSES,
            -result["latest_model_year"],
            result["manufacturer"],
            result["epa_model"],
        ),
    )


def markdown_report(results: list[dict], source_rows: int, source_path: Path) -> str:
    statuses = defaultdict(int)
    manufacturers = defaultdict(lambda: defaultdict(int))
    for result in results:
        statuses[result["match_status"]] += 1
        manufacturers[result["manufacturer"]]["total"] += 1
        manufacturers[result["manufacturer"]][result["match_status"]] += 1
        if result["probable_database_matches"]:
            manufacturers[result["manufacturer"]]["probable"] += 1

    mapped_total = sum(
        1 for result in results if result["mapped_database_brands"]
    )
    mapped_matches = sum(
        1
        for result in results
        if result["mapped_database_brands"]
        and result["match_status"] in REPRESENTED_STATUSES
    )
    recent_unmatched = [
        result
        for result in results
        if result["match_status"] not in REPRESENTED_STATUSES
        and result["latest_model_year"] >= 2024
    ]
    constant_speed_models = [
        result
        for result in results
        if "Constant Speed" in result["engine_operations"]
    ]
    variable_speed_only_models = [
        result
        for result in results
        if "Constant Speed" not in result["engine_operations"]
        and "Variable Speed" in result["engine_operations"]
    ]
    generator_priority = [
        result
        for result in recent_unmatched
        if result["mapped_database_brands"]
        and "Constant Speed" in result["engine_operations"]
    ]
    next_tier_priority = [
        result
        for result in results
        if result["match_status"] not in REPRESENTED_STATUSES
        and 2020 <= result["latest_model_year"] <= 2023
        and result["mapped_database_brands"]
        and "Constant Speed" in result["engine_operations"]
    ]
    legacy_2019_priority = [
        result
        for result in results
        if result["match_status"] not in REPRESENTED_STATUSES
        and result["latest_model_year"] == 2019
        and result["mapped_database_brands"]
        and "Constant Speed" in result["engine_operations"]
    ]
    legacy_2018_priority = [
        result
        for result in results
        if result["match_status"] not in REPRESENTED_STATUSES
        and result["latest_model_year"] == 2018
        and result["mapped_database_brands"]
        and "Constant Speed" in result["engine_operations"]
    ]
    legacy_2017_priority = [
        result
        for result in results
        if result["match_status"] not in REPRESENTED_STATUSES
        and result["latest_model_year"] == 2017
        and result["mapped_database_brands"]
        and "Constant Speed" in result["engine_operations"]
    ]
    legacy_2016_priority = [
        result
        for result in results
        if result["match_status"] not in REPRESENTED_STATUSES
        and result["latest_model_year"] == 2016
        and result["mapped_database_brands"]
        and "Constant Speed" in result["engine_operations"]
    ]
    legacy_2015_priority = [
        result
        for result in results
        if result["match_status"] not in REPRESENTED_STATUSES
        and result["latest_model_year"] == 2015
        and result["mapped_database_brands"]
        and "Constant Speed" in result["engine_operations"]
    ]
    legacy_2014_priority = [
        result
        for result in results
        if result["match_status"] not in REPRESENTED_STATUSES
        and result["latest_model_year"] == 2014
        and result["mapped_database_brands"]
        and "Constant Speed" in result["engine_operations"]
    ]
    legacy_2013_priority = [
        result
        for result in results
        if result["match_status"] not in REPRESENTED_STATUSES
        and result["latest_model_year"] == 2013
        and result["mapped_database_brands"]
        and "Constant Speed" in result["engine_operations"]
    ]
    legacy_2012_priority = [
        result
        for result in results
        if result["match_status"] not in REPRESENTED_STATUSES
        and result["latest_model_year"] == 2012
        and result["mapped_database_brands"]
        and "Constant Speed" in result["engine_operations"]
    ]
    legacy_2011_priority = [
        result
        for result in results
        if result["match_status"] not in REPRESENTED_STATUSES
        and result["latest_model_year"] == 2011
        and result["mapped_database_brands"]
        and "Constant Speed" in result["engine_operations"]
    ]
    generator_priority_brands = defaultdict(int)
    for result in generator_priority:
        for brand in result["mapped_database_brands"]:
            generator_priority_brands[brand] += 1
    exact_primary_1800 = sum(
        1
        for result in results
        if result["match_status"] == "exact_brand_match"
        and any(
            match["rpm_rated"] == 1800
            for match in result["database_matches"]
        )
    )

    lines = [
        "# EPA Nonroad CI 1800 RPM Engine Coverage",
        "",
        f"Source workbook: `{source_path.name}`",
        "",
        "## Method",
        "",
        "- Kept only `Model Info` rows where `Rated Speed (RPM)` is exactly 1800.",
        "- Joined manufacturer, certificate, tier, compliance standard and fuel from `Family Info` using model year and engine family.",
        "- Deduplicated recurring annual certifications by EPA manufacturer plus normalized engine model.",
        "- Counted a database model as present only when its normalized model and verified manufacturer-to-brand mapping both matched.",
        "- Counted a redundant leading database brand as represented only when the remaining normalized model had at least five characters.",
        "- Counted slash-suffixed certification trims as represented only when the explicit base model before `/` matched the verified database brand.",
        "- Counted non-slash certification trims only through reviewed manufacturer, brand and suffix-pattern rules.",
        "- Counted commercial family variants only through reviewed manufacturer, brand and prefix rules; short families also require an EPA emissions label on the matched page.",
        "- Counted non-literal certification aliases only from the reviewed `CERTIFICATION_ALIASES` map.",
        "- Counted certification groups only when every commercial model in the reviewed `CERTIFICATION_ALIAS_GROUPS` map was present.",
        "- Exact model matches under another brand and other similar suffix variants remain review items.",
        "",
        "## Summary",
        "",
        f"- 1800 RPM source rows: **{source_rows:,}**",
        f"- Distinct EPA manufacturer/model combinations: **{len(results):,}**",
        f"- Exact manufacturer/brand matches: **{statuses['exact_brand_match']:,}**",
        f"- Matches after removing a redundant database brand prefix: **{statuses['brand_prefix_match']:,}**",
        f"- Slash-suffixed certification trims represented by a verified base model: **{statuses['base_brand_match']:,}**",
        f"- Reviewed manufacturer certification trims: **{statuses['certification_trim_match']:,}**",
        f"- Reviewed certification aliases: **{statuses['certification_alias_match']:,}**",
        f"- Fully represented certification groups: **{statuses['certification_alias_group_match']:,}**",
        f"- Verified commercial family matches: **{statuses['family_brand_match']:,}**",
        f"- Exact matches whose database page uses 1800 as its primary RPM: **{exact_primary_1800:,}**",
        f"- Exact model under another database brand: **{statuses['exact_model_other_brand']:,}**",
        f"- Not represented after reviewed matching rules: **{statuses['not_found']:,}**",
        f"- Models from mapped manufacturers: **{mapped_total:,}**",
        f"- Represented coverage within mapped manufacturers: **{mapped_matches / mapped_total:.1%}**",
        f"- Unmatched models with a 2024+ certification: **{len(recent_unmatched):,}**",
        f"- Models with at least one constant-speed certification: **{len(constant_speed_models):,}**",
        f"- Variable-speed-only models retained for reference: **{len(variable_speed_only_models):,}**",
        f"- Generator-priority review queue (2024+, mapped brand, constant speed): **{len(generator_priority):,}**",
        f"- Next-tier review queue (2020–2023, mapped brand, constant speed): **{len(next_tier_priority):,}**",
        f"- Legacy 2019 review queue (mapped brand, constant speed): **{len(legacy_2019_priority):,}**",
        f"- Legacy 2018 review queue (mapped brand, constant speed): **{len(legacy_2018_priority):,}**",
        f"- Legacy 2017 review queue (mapped brand, constant speed): **{len(legacy_2017_priority):,}**",
        f"- Legacy 2016 review queue (mapped brand, constant speed): **{len(legacy_2016_priority):,}**",
        f"- Legacy 2015 review queue (mapped brand, constant speed): **{len(legacy_2015_priority):,}**",
        f"- Legacy 2014 review queue (mapped brand, constant speed): **{len(legacy_2014_priority):,}**",
        f"- Legacy 2013 review queue (mapped brand, constant speed): **{len(legacy_2013_priority):,}**",
        f"- Legacy 2012 review queue (mapped brand, constant speed): **{len(legacy_2012_priority):,}**",
        f"- Legacy 2011 review queue (mapped brand, constant speed): **{len(legacy_2011_priority):,}**",
        "",
        "The primary RPM field does not prove that a page lacks 60 Hz ratings; many catalog pages use "
        "1500 RPM as the primary value while storing separate 60 Hz fields. Those pages need a second "
        "rating-level comparison before any RPM correction.",
        "",
        "## Manufacturer Mapping Notes",
        "",
        "- `Discovery Energy, LLC.` is compared with the existing `Kohler` brand. Rehlko's official "
        "[engine warranty page](https://www.engines.rehlko.com/warranty) identifies Discovery Energy "
        "as the responsible company and states that Kohler Engines is now Rehlko.",
        "- `HD Construction Equipment Co., Ltd.` is compared with the existing `Hyundai` brand. Its "
        "[official network page](https://www.hd-ce.com/en/network) lists the current company and its "
        "engine production and R&D operations.",
        "- `Caterpillar Inc.` is also compared with `Perkins` for reviewed family-prefix matches. "
        "Official Perkins product documentation identifies commercial models such as "
        "`1706J-E93TA` and `2406J-E13TA` behind the shorter EPA family names.",
        "- `Cummins Inc.` exact generator-drive pages retain commercial `G` and `NR2` suffixes. "
        "Three reviewed aliases map generic `-G` certification names to existing QSK78, QST30 "
        "and QSX15 family pages where displacement, certification tiers and 1800 RPM operation "
        "already agree.",
        "- `Mitsubishi Heavy Industries Engine & Turbocharger, Ltd.` EPA records shorten six "
        "Tier 2 generator models by omitting their `Y2` application codes. Reviewed aliases map "
        "only to existing Y2PTAW 60 Hz pages whose EPA engine code, displacement and certified "
        "1800 RPM power node agree.",
        "- `FPT Industrial S.p.A.` uses internal EPA certification codes for its Tier 3 engines. "
        "Twelve reviewed aliases are mapped to N45, N67 and Cursor 9 commercial pages only where "
        "displacement, certification family and published power node agree with FPT's official "
        "power-generation brochure.",
        "- `Liebherr Machines Bulle SA` is compared with both `Liebherr` and `Kohler`. Liebherr's "
        "official co-development announcement identifies the six KD commercial engine families "
        "manufactured for Kohler generator sets.",
        "- `Kubota Corporation` EPA model names use certification suffixes such as `-EF` and "
        "`-ET`, while Kubota's public generator catalog uses commercial `E4-BG` and `E3-BG` "
        "names. Sixteen reviewed aliases require matching displacement, aspiration, emissions "
        "tier and Kubota-published 1800 RPM output.",
        "- `Perkins Engines Co Ltd` EPA records omit the generator-drive `G` suffix and may append "
        "the shared Caterpillar base-engine name in parentheses. Seventeen reviewed aliases map those "
        "records only to Perkins commercial pages with matching family, displacement, emissions "
        "tier and manufacturer-published 1800 RPM power node.",
        "- `Rolls-Royce Solutions America Inc` is represented under the `MTU` database brand. "
        "Exact 60 Hz commercial model pages retain MTU's `S`, `3B` and `3D` application suffixes "
        "and use the latest 1800 RPM power node in the EPA workbook. Fifteen reviewed aliases cover "
        "EPA certification names whose only difference is the presence or absence of the verified "
        "`3B` or `3D` suffix, supplemented by public MTU gendrive specifications and operating "
        "instructions where available.",
        "- `Yanmar Power Technology Co., Ltd.` uses several EPA certification configuration "
        "names that differ from its TNV generator product names. Seventeen reviewed aliases map only "
        "where displacement, aspiration, emissions tier and the certified power node agree with "
        "Yanmar's official generator and industrial-engine documentation.",
        "",
        "## Generator-Priority Gaps by Brand",
        "",
        "| Database brand | 2024+ constant-speed models without represented match |",
        "|---|---:|",
    ]

    for brand, count in sorted(
        generator_priority_brands.items(), key=lambda item: (-item[1], item[0])
    ):
        lines.append(f"| {brand} | {count} |")
    if not generator_priority_brands:
        lines.append("| None | 0 |")

    lines.extend(
        [
        "",
        "## Manufacturer Coverage",
        "",
        "| EPA manufacturer | Database brand | EPA models | Exact | Brand prefix | Base trims | Cert. trims | Cert. aliases | Families | Other-brand exact | Not found | Probable |",
        "|---|---|---:|---:|---:|---:|---:|---:|---:|---:|---:|---:|",
        ]
    )

    for manufacturer, counts in sorted(
        manufacturers.items(), key=lambda item: (-item[1]["total"], item[0])
    ):
        brands = ", ".join(sorted(MANUFACTURER_BRANDS.get(manufacturer, set()))) or "Unmapped"
        lines.append(
            f"| {manufacturer} | {brands} | {counts['total']} | "
            f"{counts['exact_brand_match']} | {counts['brand_prefix_match']} | "
            f"{counts['base_brand_match']} | "
            f"{counts['certification_trim_match']} | "
            f"{counts['certification_alias_match']} | "
            f"{counts['family_brand_match']} | "
            f"{counts['exact_model_other_brand']} | "
            f"{counts['not_found']} | {counts['probable']} |"
        )

    lines.extend(
        [
            "",
            "## Priority Review",
            "",
            "These are recent constant-speed EPA-certified models from mapped manufacturers that "
            "were not found as exact, brand-prefix, base-trim, reviewed certification-trim, "
            "reviewed certification-alias or "
            "verified commercial-family matches. "
            "The full records, including "
            "variable-speed-only models and candidate aliases, are in "
            "`epa-1800rpm-model-match.json`.",
            "",
            "| Latest year | Manufacturer | EPA model | Tier | Power kW | Probable database model |",
            "|---:|---|---|---|---:|---|",
        ]
    )

    for result in sorted(
        generator_priority,
        key=lambda item: (
            -item["latest_model_year"],
            item["manufacturer"],
            item["epa_model"],
        ),
    )[:150]:
        tier = ", ".join(str(value) for value in result["applicable_tiers"])
        power = (
            str(result["rated_power_kw_max"])
            if result["rated_power_kw_max"] is not None
            else ""
        )
        probable = ""
        if result["probable_database_matches"]:
            candidate = result["probable_database_matches"][0]
            probable = f"{candidate['brand']} {candidate['model']} ({candidate['similarity']:.3f})"
        lines.append(
            f"| {result['latest_model_year']} | {result['manufacturer']} | "
            f"{result['epa_model']} | {tier} | {power} | {probable} |"
        )
    if not generator_priority:
        lines.append("| - | None | None | - | - | - |")

    lines.extend(
        [
            "",
            "## Interpretation",
            "",
            "- EPA certification records identify certified engine configurations, not generator-set electrical ratings.",
            "- A shared base model can have different certification families, power ratings, aftertreatment and model years.",
            "- Candidate additions should be validated against the EPA certificate and a manufacturer datasheet before database insertion.",
            "",
        ]
    )
    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--epa-xlsx", required=True, type=Path)
    parser.add_argument("--engines-json", required=True, type=Path)
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("reports/epa-certification"),
    )
    args = parser.parse_args()

    workbook = openpyxl.load_workbook(
        args.epa_xlsx, read_only=True, data_only=True
    )
    families = read_family_info(workbook)
    models, source_rows = read_1800_rpm_models(workbook, families)
    engines = json.loads(args.engines_json.read_text())
    results = match_models(models, engines)

    args.output_dir.mkdir(parents=True, exist_ok=True)
    json_path = args.output_dir / "epa-1800rpm-model-match.json"
    report_path = args.output_dir / "epa-1800rpm-summary.md"
    json_path.write_text(json.dumps(results, indent=2, ensure_ascii=False) + "\n")
    report_path.write_text(
        markdown_report(results, source_rows, args.epa_xlsx) + "\n"
    )

    print(f"Wrote {report_path}")
    print(f"Wrote {json_path}")


if __name__ == "__main__":
    main()
