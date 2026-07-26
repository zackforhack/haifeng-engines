#!/usr/bin/env python3
"""Analyze exact-1800 RPM EPA spark-ignited certification coverage.

The EPA SI workbook is family/test-level data and does not contain an engine
model column. This analyzer therefore reports annual families and carryover
lineages, then produces configuration candidates rather than claiming exact
model coverage.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl


MANUFACTURER_BRANDS = {
    "2G Heek GmbH": {"2G"},
    "AB Volvo Penta": {"Volvo Penta"},
    "Arrow Engine Company": {"Arrow"},
    # Baseline's NG 400/NexGen 400 uses the same 21.9 L V12,
    # 580 hp calibration as PSI's EPA-certified 22 L platform.
    "Baseline Energy Services, LP": {"PSI"},
    "Caterpillar Inc.": {"Caterpillar"},
    "Cummins Inc.": {"Cummins"},
    "Deutz AG": {"Deutz"},
    "Discovery Energy, LLC.": {"Kohler"},
    "DNGV Co., Ltd.": {"DNGV"},
    # ENER-G's published natural-gas CHP range identifies MAN engines,
    # including the E3262 LE202 used in its 530M package.
    "ENER-G Rudox LLC": {"ENER-G", "MAN", "Mitsubishi"},
    # Enchanted Rock's 2026 X22 certification is documented with the
    # HDI GX22 engine, represented by the Mesa/HDI GX22 catalog page.
    "Enchanted Rock, LLC": {"Mesa"},
    "Generac Power Systems, Inc.": {"Generac"},
    "Guascor Energy S.A.U.": {"Guascor"},
    "IMPCO Technologies, Inc.": {"IMPCO"},
    "INNIO Jenbacher GmbH & CO OG": {"Jenbacher"},
    "INNIO Waukesha Gas Engines Inc.": {"Waukesha"},
    "KEM Equipment, Inc.": {"KEM"},
    "Kubota Corporation": {"Kubota"},
    "Mesa Natural Gas Solutions, LLC": {"Mesa"},
    "Origin Engines": {"Origin Engines"},
    "Power Solutions International": {"PSI"},
    "Power Solutions International, Inc.": {"PSI"},
    "Rolls-Royce Solutions America Inc": {"MTU"},
    # Scale's 2026 26.0 L V12 family code ends in MAN and its 301 kW
    # node matches MAN's 25.8 L E3262 E302 300 kW / 1800 RPM rating.
    "Scale Microgrid Solutions Operating, LLC": {"MAN"},
    # Tecogen's lineage spans its legacy GM-supplied TecoDrive 7400 and
    # the newer Origin 8.0 L used in InVerde e+ and TecoPower products.
    "Tecogen": {"Tecogen", "Origin Engines"},
    # Tennessee Propulsion Products certified the Isuzu 4HV1
    # alternative-fuel industrial platform under manufacturer code IND.
    "Tennessee Propulsion Products, LLC": {"Isuzu"},
    # Weichai America certifies some PSI-branded gaseous engines. The
    # PWCAB16.7GTA lineage, for example, is the PSI 17L platform.
    "Weichai America Corporation": {"Weichai", "PSI"},
    "Yanmar Power Technology Co., Ltd.": {"Yanmar"},
    "Zenith Power Products": {"Zenith Power Products"},
}

STATIONARY_MODES = {
    "Part 60 Subpart JJJJ Table 1",
    "Stationary",
    "Stationary Part 1048",
}

RESEARCH_EXCEPTION_NOTES = {
    (
        "INNIO Waukesha Gas Engines Inc.",
        "TDRSB24.0LTP",
    ): (
        "INNIO's official Waukesha power-rating guide publishes current "
        "24 L H24 generator-drive ratings of 400 or 440 kWb at 1800 RPM. "
        "The EPA lineage reaches 550 kW and includes mobile and stationary "
        "test modes, so the catalog H24 models cannot be claimed as this "
        "lineage without a matching certificate or calibration sheet."
    ),
    (
        "Generac Power Systems, Inc.",
        "GGNXB08.92C6",
    ): (
        "The closest catalog configuration is Generac's naturally aspirated "
        "8.9 L SG080, but its published 91 kWm rating is 31.5% above the EPA "
        "family's 69.19 kW node. Searches of Generac documentation and "
        "archived equipment records did not identify a commercial model for "
        "GGNXB08.92C6, so the family remains quarantined."
    ),
    (
        "Power Solutions International, Inc.",
        "SPSIB04.3TNP",
    ): (
        "PSI's current stationary owner manual acknowledges a 4.3 L V6 "
        "turbo configuration, but its public power-systems catalog and "
        "downloadable 4.3 L generator sheet publish only the naturally "
        "aspirated 61 kW model. This 2025-2026 EPA lineage is explicitly "
        "turbocharged and reaches 75 and 89.6 kW, so it remains quarantined "
        "until PSI publishes a commercial model and stationary rating sheet."
    ),
    (
        "Rolls-Royce Solutions America Inc",
        "DMDDB06.8GBX",
    ): (
        "The cataloged MTU 10V0068 GS100 is turbocharged, while EPA identifies "
        "this 6.8 L V10 lineage as naturally aspirated. A historical MTU "
        "specification or certificate is required before adding the model."
    ),
    (
        "Rolls-Royce Solutions America Inc",
        "EMDDB06.8GBX",
    ): (
        "The cataloged MTU 10V0068 GS100 is turbocharged, while EPA identifies "
        "this 2014 6.8 L V10 family as naturally aspirated. A historical MTU "
        "specification or certificate is required before adding the model."
    ),
    (
        "Gridiron, LLC",
        "PGRIB02.2CHP",
    ): (
        "DOE CHP eCatalog identifies Gridiron PowerPlant H24 and HA65 packages "
        "and says they use purpose-built gaseous engines, but it does not "
        "publish an engine model or enough configuration data to crosswalk "
        "this 2.2 L family."
    ),
    (
        "Zenith Power Products",
        "EZPPB12.9TAC",
    ): (
        "Zenith's published TA6120 is an 11.8 L engine and its EPA lineage "
        "starts in 2017. It must not be used as a substitute for this distinct "
        "2014-2015 12.9 L family without a historical certificate or manual."
    ),
}

# Matching corrections are limited to conflicts resolved by primary
# manufacturer documentation. Raw EPA values remain in the report JSON.
SOURCE_CONFIGURATION_OVERRIDES = {
    (
        "ENER-G Rudox LLC",
        "FRDXB65.5MGS",
    ): {
        "displacements_l": [65.37],
        "reason": (
            "EPA reports 49.1 L for a V16 family coded 65.5; the Rudox "
            "ERM1000GS datasheet identifies its Mitsubishi GS16R-PTK, "
            "and MHI publishes that V16 engine at 65.37 L."
        ),
    },
    (
        "Cummins Inc.",
        "HCEXB38.0AAA",
    ): {
        "cylinders": [12],
        "reason": (
            "EPA reports 16 cylinders; Cummins identifies the 37.7 L "
            "KTA38GC platform as a V12."
        ),
    },
    (
        "Power Solutions International",
        "APWRB18.3NGP",
    ): {
        "displacements_l": [18.3],
        "reason": (
            "EPA reports 14.6 L; the family code and official PSI 18L "
            "catalog identify an 18.3 L V10 at approximately 422 kWm."
        ),
    },
    (
        "Power Solutions International, Inc.",
        "APWRB18.3NGP",
    ): {
        "displacements_l": [18.3],
        "reason": (
            "EPA reports 14.6 L; the family code and official PSI 18L "
            "catalog identify an 18.3 L V10 at approximately 422 kWm."
        ),
    },
    (
        "Power Solutions International, Inc.",
        "BPWRB18.3NGP",
    ): {
        "displacements_l": [18.3],
        "reason": (
            "EPA reports 14.6 L; the family code and official PSI 18L "
            "catalog identify an 18.3 L V10 at approximately 422 kWm."
        ),
    },
    (
        "Power Solutions International, Inc.",
        "DPWRB18.3NGP",
    ): {
        "displacements_l": [18.3],
        "reason": (
            "EPA reports 14.6 L; the family code and official PSI 18L "
            "catalog identify an 18.3 L V10 at 422 kWm."
        ),
    },
    (
        "Power Solutions International, Inc.",
        "EPWRB18.3NGP",
    ): {
        "displacements_l": [18.3],
        "reason": (
            "EPA reports 14.6 L; the family code and official PSI 18L "
            "catalog identify an 18.3 L V10 at 422 kWm."
        ),
    },
    (
        "Power Solutions International, Inc.",
        "FPWRB18.3NGP",
    ): {
        "displacements_l": [18.3],
        "reason": (
            "EPA reports 14.6 L; the family code and official PSI 18L "
            "catalog identify an 18.3 L V10 at 422 kWm."
        ),
    },
    (
        "Weichai America Corporation",
        "PWCAB16.7GTA",
    ): {
        "displacements_l": [16.7],
        "reason": (
            "EPA reports 12.5 L; the family code and official PSI 17L "
            "catalog identify a 16.7 L V8 at 460 kWm."
        ),
    },
    (
        "Generac Power Systems, Inc.",
        "JGNXB02.42N1",
    ): {
        "displacements_l": [2.4],
        "reason": (
            "EPA reports 6.8 L with four cylinders; Generac identifies "
            "the QT025A family as a 2.4 L inline-four and the family "
            "code itself contains 02.4."
        ),
    },
}

VERIFIED_LINEAGE_MODEL_CROSSWALKS = {
    (
        "Generac Power Systems, Inc.",
        "EGNXB08.92C5",
    ): {
        "slug": "generac-mgg100m",
        "reason": (
            "Multiple independent equipment records identify MGG100M "
            "packages with a Generac 8.9 L V8 nameplate carrying EPA family "
            "EGNXB08.92C5, 150 hp and 1800 RPM."
        ),
    },
    (
        "Cummins Inc.",
        "ECEXB02.4AAA",
    ): {
        "slug": "cummins-c20n6",
        "reason": (
            "Cummins Power Generation's EPA compliance statement identifies "
            "the C20N6, QSJ2.4 engine and ECEXB02.4AAA family directly."
        ),
    },
    (
        "Arrow Engine Company",
        "MARWB05.4A54",
    ): {
        "slug": "arrow-a54e",
        "reason": (
            "Arrow's A54E specification and EPA certification scope identify "
            "the same 5.4 L inline-six naturally aspirated platform."
        ),
    },
    (
        "Kubota Corporation",
        "DKBXB02.52FM",
    ): {
        "slug": "kubota-wg2503-ln-e3",
        "reason": (
            "Kubota's WG2503-LN-E3 documentation and the EPA lineage identify "
            "the same 2.5 L inline-four naturally aspirated stationary platform."
        ),
    },
    (
        "Origin Engines",
        "RORGB03.6PTA",
    ): {
        "slug": "origin-engines-3-6l-turbo",
        "reason": (
            "Origin's 3.6 L turbo generator configuration and EPA annual data "
            "identify the same 3.6 L inline-four turbocharged platform."
        ),
    },
}


def normalize(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).upper()
    return re.sub(r"[^A-Z0-9]+", "", text)


def parse_displacements(value: object) -> list[float]:
    return [
        float(match)
        for match in re.findall(r"(\d+(?:\.\d+)?)\s*L", str(value or ""), re.I)
    ]


def sorted_values(values: set[object]) -> list[object]:
    return sorted(value for value in values if value not in (None, "", "-"))


def read_workbook(path: Path) -> tuple[list[dict], int]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    index = {header: position for position, header in enumerate(headers)}

    all_family_parents: dict[tuple[str, str], str] = {}
    annual_families: dict[tuple[str, str, int], dict] = {}
    source_rows = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        manufacturer = str(row[index["Manufacturer"]] or "").strip()
        family = str(row[index["Engine Family"]] or "").strip()
        year_value = row[index["Model Year"]]
        if not manufacturer or not family or year_value is None:
            continue

        year = int(year_value)
        carryover = str(row[index["Carryover Engine Family"]] or "").strip()
        if carryover and carryover != "-":
            all_family_parents[(manufacturer, family)] = carryover

        speed = row[index["Max Engine Test Speed (RPM)"]]
        if not isinstance(speed, (int, float)) or abs(speed - 1800) > 0.01:
            continue

        source_rows += 1
        key = (manufacturer, family, year)
        if key not in annual_families:
            annual_families[key] = {
                "manufacturer": manufacturer,
                "manufacturer_code": row[index["Manufacturer Code"]],
                "engine_family": family,
                "model_year": year,
                "carryover_families": set(),
                "cylinders": set(),
                "arrangements": set(),
                "fuel_categories": set(),
                "test_fuels": set(),
                "test_cycles": set(),
                "mobile_stationary_tests": set(),
                "displacement_texts": set(),
                "displacements_l": set(),
                "max_powers_kw": set(),
                "aspiration_methods": set(),
                "cert_test_numbers": set(),
            }

        record = annual_families[key]
        record["carryover_families"].add(carryover)
        record["cylinders"].add(row[index["Cylinders"]])
        record["arrangements"].add(row[index["Cylinder Arrangement"]])
        record["fuel_categories"].add(row[index["Engine Fuel Category"]])
        record["test_fuels"].add(row[index["Test Fuel"]])
        record["test_cycles"].add(row[index["Test Cycle"]])
        record["mobile_stationary_tests"].add(
            row[index["Mobile or Stationary Test"]]
        )
        displacement_text = row[index["Displacement"]]
        record["displacement_texts"].add(displacement_text)
        record["displacements_l"].update(parse_displacements(displacement_text))
        record["max_powers_kw"].add(row[index["Max Power (kW)"]])
        record["aspiration_methods"].add(row[index["Aspiration Method"]])
        record["cert_test_numbers"].add(row[index["Cert test number"]])

    def root_family(manufacturer: str, family: str) -> str:
        seen = set()
        current = family
        while (manufacturer, current) in all_family_parents and current not in seen:
            seen.add(current)
            current = all_family_parents[(manufacturer, current)]
        return current

    results = []
    for record in annual_families.values():
        modes = sorted_values(record["mobile_stationary_tests"])
        results.append(
            {
                "manufacturer": record["manufacturer"],
                "manufacturer_code": record["manufacturer_code"],
                "engine_family": record["engine_family"],
                "lineage_root_family": root_family(
                    record["manufacturer"], record["engine_family"]
                ),
                "model_year": record["model_year"],
                "carryover_families": sorted_values(record["carryover_families"]),
                "cylinders": sorted_values(record["cylinders"]),
                "arrangements": sorted_values(record["arrangements"]),
                "fuel_categories": sorted_values(record["fuel_categories"]),
                "test_fuels": sorted_values(record["test_fuels"]),
                "test_cycles": sorted_values(record["test_cycles"]),
                "mobile_stationary_tests": modes,
                "has_stationary_test": any(mode in STATIONARY_MODES for mode in modes),
                "displacement_texts": sorted_values(record["displacement_texts"]),
                "displacements_l": sorted_values(record["displacements_l"]),
                "max_powers_kw": sorted_values(record["max_powers_kw"]),
                "aspiration_methods": sorted_values(record["aspiration_methods"]),
                "cert_test_numbers": sorted_values(record["cert_test_numbers"]),
            }
        )

    return results, source_rows


def aggregate_lineages(annual_families: list[dict]) -> list[dict]:
    grouped: dict[tuple[str, str], dict] = {}
    set_fields = [
        "annual_engine_families",
        "model_years",
        "cylinders",
        "arrangements",
        "fuel_categories",
        "test_fuels",
        "test_cycles",
        "mobile_stationary_tests",
        "displacement_texts",
        "displacements_l",
        "max_powers_kw",
        "aspiration_methods",
        "cert_test_numbers",
    ]

    for family in annual_families:
        key = (family["manufacturer"], family["lineage_root_family"])
        if key not in grouped:
            grouped[key] = {
                "manufacturer": family["manufacturer"],
                "manufacturer_code": family["manufacturer_code"],
                "lineage_root_family": family["lineage_root_family"],
                "has_stationary_test": False,
                **{field: set() for field in set_fields},
            }

        lineage = grouped[key]
        lineage["annual_engine_families"].add(family["engine_family"])
        lineage["model_years"].add(family["model_year"])
        lineage["has_stationary_test"] |= family["has_stationary_test"]
        for field in set_fields[2:]:
            lineage[field].update(family[field])

    results = []
    for lineage in grouped.values():
        results.append(
            {
                **{
                    key: value
                    for key, value in lineage.items()
                    if not isinstance(value, set)
                },
                **{
                    field: sorted_values(lineage[field])
                    for field in set_fields
                },
            }
        )
    return results


def database_power_values(engine: dict) -> list[float]:
    fields = [
        "power_kw",
        "prime_power_kw_60hz",
        "standby_power_kw_60hz",
    ]
    return [
        float(engine[field])
        for field in fields
        if isinstance(engine.get(field), (int, float))
    ]


def has_1800_capability(engine: dict) -> bool:
    return engine.get("rpm_rated") == 1800 or any(
        engine.get(field) is not None
        for field in (
            "prime_power_kw_60hz",
            "standby_power_kw_60hz",
            "prime_power_kwe_60hz",
            "standby_power_kwe_60hz",
        )
    )


def is_gas_engine(engine: dict) -> bool:
    return str(engine.get("fuel_type") or "") not in {"Diesel", "Diesel/HVO"}


def aspiration_categories(values: list[object]) -> set[str]:
    categories = set()
    for value in values:
        normalized = normalize(value)
        if "NATURALLYASPIRATED" in normalized:
            categories.add("naturally_aspirated")
        if "TURBO" in normalized or "SUPERCHARG" in normalized:
            categories.add("forced_induction")
    return categories


def engine_aspiration_category(engine: dict) -> str | None:
    categories = aspiration_categories([engine.get("configuration")])
    return next(iter(categories)) if len(categories) == 1 else None


def candidate_matches(lineage: dict, engines: list[dict]) -> list[dict]:
    expected_brands = MANUFACTURER_BRANDS.get(lineage["manufacturer"], set())
    override = SOURCE_CONFIGURATION_OVERRIDES.get(
        (lineage["manufacturer"], lineage["lineage_root_family"]),
        {},
    )
    displacements = override.get(
        "displacements_l",
        lineage["displacements_l"],
    )
    cylinders = {
        int(value)
        for value in override.get("cylinders", lineage["cylinders"])
        if isinstance(value, (int, float))
    }
    powers = {
        float(value)
        for value in lineage["max_powers_kw"]
        if isinstance(value, (int, float))
    }
    source_aspirations = aspiration_categories(lineage["aspiration_methods"])
    required_aspiration = (
        next(iter(source_aspirations)) if len(source_aspirations) == 1 else None
    )
    verified_crosswalk = VERIFIED_LINEAGE_MODEL_CROSSWALKS.get(
        (lineage["manufacturer"], lineage["lineage_root_family"])
    )
    candidates = []

    for engine in engines:
        if engine.get("brand") not in expected_brands:
            continue
        if not is_gas_engine(engine) or not has_1800_capability(engine):
            continue
        if cylinders and engine.get("cylinders") not in cylinders:
            continue
        engine_aspiration = engine_aspiration_category(engine)
        if (
            required_aspiration
            and engine_aspiration
            and engine_aspiration != required_aspiration
        ):
            continue

        engine_displacement = engine.get("displacement_l")
        if not isinstance(engine_displacement, (int, float)) or not displacements:
            continue
        displacement_delta = min(
            abs(float(engine_displacement) - float(value))
            for value in displacements
        )
        displacement_tolerance = max(0.1, float(engine_displacement) * 0.03)
        if displacement_delta > displacement_tolerance:
            continue

        engine_powers = database_power_values(engine)
        power_delta_ratio = None
        if powers and engine_powers:
            power_delta_ratio = min(
                abs(engine_power - source_power) / max(source_power, 1)
                for engine_power in engine_powers
                for source_power in powers
            )

        candidates.append(
            {
                "brand": engine["brand"],
                "model": engine["model"],
                "slug": engine["slug"],
                "fuel_type": engine.get("fuel_type"),
                "displacement_l": engine_displacement,
                "cylinders": engine.get("cylinders"),
                "rpm_rated": engine.get("rpm_rated"),
                "source_aspiration": required_aspiration,
                "database_aspiration": engine_aspiration,
                "database_power_kw": engine_powers,
                "verified_lineage_crosswalk": bool(
                    verified_crosswalk
                    and verified_crosswalk["slug"] == engine["slug"]
                ),
                "source_power_delta_ratio": (
                    round(power_delta_ratio, 3)
                    if power_delta_ratio is not None
                    else None
                ),
            }
        )

    return sorted(
        candidates,
        key=lambda candidate: (
            candidate["source_power_delta_ratio"] is None,
            candidate["source_power_delta_ratio"]
            if candidate["source_power_delta_ratio"] is not None
            else 999,
            candidate["slug"],
        ),
    )[:12]


def classify_lineages(lineages: list[dict], engines: list[dict]) -> list[dict]:
    results = []
    for lineage in lineages:
        expected_brands = sorted(
            MANUFACTURER_BRANDS.get(lineage["manufacturer"], set())
        )
        candidates = candidate_matches(lineage, engines)
        strong_candidates = [
            candidate
            for candidate in candidates
            if candidate["verified_lineage_crosswalk"]
            or (
                candidate["source_power_delta_ratio"] is not None
                and candidate["source_power_delta_ratio"] <= 0.15
            )
        ]
        if strong_candidates:
            status = "strong_configuration_candidate"
        elif (
            lineage["manufacturer"],
            lineage["lineage_root_family"],
        ) in RESEARCH_EXCEPTION_NOTES:
            status = "researched_configuration_exception"
        elif candidates:
            status = "configuration_candidate"
        elif expected_brands:
            status = "mapped_manufacturer_no_configuration_candidate"
        else:
            status = "unmapped_manufacturer"

        results.append(
            {
                **lineage,
                "mapped_database_brands": expected_brands,
                "matching_override": SOURCE_CONFIGURATION_OVERRIDES.get(
                    (
                        lineage["manufacturer"],
                        lineage["lineage_root_family"],
                    )
                ),
                "verified_lineage_crosswalk": VERIFIED_LINEAGE_MODEL_CROSSWALKS.get(
                    (
                        lineage["manufacturer"],
                        lineage["lineage_root_family"],
                    )
                ),
                "coverage_status": status,
                "database_candidates": candidates,
            }
        )

    return sorted(
        results,
        key=lambda result: (
            result["coverage_status"] != "unmapped_manufacturer",
            result["coverage_status"]
            != "mapped_manufacturer_no_configuration_candidate",
            -max(result["model_years"]),
            result["manufacturer"],
            result["lineage_root_family"],
        ),
    )


def markdown_report(
    annual_families: list[dict],
    lineages: list[dict],
    source_rows: int,
    source_path: Path,
    source_sha256: str,
    catalog_engine_count: int,
    catalog_sha256: str,
) -> str:
    statuses = Counter(lineage["coverage_status"] for lineage in lineages)
    manufacturers = defaultdict(lambda: Counter())
    for lineage in lineages:
        bucket = manufacturers[lineage["manufacturer"]]
        bucket["lineages"] += 1
        bucket[lineage["coverage_status"]] += 1
        bucket["annual_families"] += len(lineage["annual_engine_families"])
        if lineage["has_stationary_test"]:
            bucket["stationary_lineages"] += 1
            bucket[f"stationary_{lineage['coverage_status']}"] += 1

    mapped_lineages = [
        lineage for lineage in lineages if lineage["mapped_database_brands"]
    ]
    stationary_lineages = [
        lineage for lineage in lineages if lineage["has_stationary_test"]
    ]
    stationary_statuses = Counter(
        lineage["coverage_status"] for lineage in stationary_lineages
    )
    remaining_stationary_gaps = [
        lineage
        for lineage in stationary_lineages
        if lineage["coverage_status"]
        in {
            "unmapped_manufacturer",
            "mapped_manufacturer_no_configuration_candidate",
            "researched_configuration_exception",
        }
    ]
    recent_stationary_gaps = [
        lineage
        for lineage in lineages
        if max(lineage["model_years"]) >= 2024
        and lineage["has_stationary_test"]
        and lineage["coverage_status"]
        in {
            "unmapped_manufacturer",
            "mapped_manufacturer_no_configuration_candidate",
        }
    ]
    recent_mobile_only_gaps = [
        lineage
        for lineage in lineages
        if max(lineage["model_years"]) >= 2024
        and not lineage["has_stationary_test"]
        and lineage["coverage_status"]
        in {
            "unmapped_manufacturer",
            "mapped_manufacturer_no_configuration_candidate",
        }
    ]
    researched_stationary_exceptions = [
        lineage
        for lineage in stationary_lineages
        if lineage["coverage_status"] == "researched_configuration_exception"
    ]
    overridden_lineages = [
        lineage for lineage in lineages if lineage["matching_override"]
    ]
    verified_crosswalks = [
        lineage for lineage in lineages if lineage["verified_lineage_crosswalk"]
    ]

    lines = [
        "# EPA Spark-Ignited 1800 RPM Coverage Baseline",
        "",
        f"Source workbook: `{source_path.name}`",
        f"Source SHA-256: `{source_sha256}`",
        f"Catalog snapshot engines: **{catalog_engine_count:,}**",
        f"Catalog snapshot SHA-256: `{catalog_sha256}`",
        "",
        "## Scope And Limitation",
        "",
        "- Kept only rows where `Max Engine Test Speed (RPM)` is exactly 1800.",
        "- Deduplicated repeated fuel and test-cycle rows into annual EPA engine families.",
        "- Followed `Carryover Engine Family` references to group recurring annual certifications into lineages.",
        "- The workbook has no engine-model column. Therefore this report does not claim exact model coverage.",
        "- Configuration candidates require a mapped manufacturer, gas-fueled catalog page, 1800/60 Hz capability, matching cylinder count and displacement within 3%.",
        "- When the EPA lineage has one unambiguous aspiration method, an explicitly conflicting catalog configuration is excluded.",
        "- A strong candidate has either a documented lineage-to-model crosswalk or a catalog mechanical-power value within 15% of an EPA maximum-power node.",
        "- Documented source-field conflicts are corrected only for matching; raw EPA values remain in the JSON output.",
        "",
        "## Summary",
        "",
        f"- Exact-1800 source rows: **{source_rows:,}**",
        f"- Annual manufacturer/family identities: **{len(annual_families):,}**",
        f"- Carryover lineages: **{len(lineages):,}**",
        f"- Manufacturers: **{len(manufacturers):,}**",
        f"- Lineages with at least one stationary test: **{len(stationary_lineages):,}**",
        f"- Lineages under mapped catalog manufacturers: **{len(mapped_lineages):,}**",
        f"- Documented source-field corrections used for matching: **{len(overridden_lineages):,}**",
        f"- Documented lineage-to-model crosswalks: **{len(verified_crosswalks):,}**",
        "",
        "### Stationary Coverage Candidates",
        "",
        f"- Strong configuration candidates: **{stationary_statuses['strong_configuration_candidate']:,}**",
        f"- Configuration candidates without close power confirmation: **{stationary_statuses['configuration_candidate']:,}**",
        f"- Researched and quarantined configuration exceptions: **{stationary_statuses['researched_configuration_exception']:,}**",
        f"- Mapped manufacturer but no configuration candidate: **{stationary_statuses['mapped_manufacturer_no_configuration_candidate']:,}**",
        f"- Unmapped manufacturer lineages: **{stationary_statuses['unmapped_manufacturer']:,}**",
        f"- Recent 2024+ stationary gaps: **{len(recent_stationary_gaps):,}**",
        f"- Recent 2024+ mobile-only gaps excluded from the queue: **{len(recent_mobile_only_gaps):,}**",
        "",
        "### All Exact-1800 Lineages",
        "",
        f"- Strong configuration candidates: **{statuses['strong_configuration_candidate']:,}**",
        f"- Other configuration candidates: **{statuses['configuration_candidate']:,}**",
        f"- Researched and quarantined configuration exceptions: **{statuses['researched_configuration_exception']:,}**",
        f"- Mapped manufacturer but no configuration candidate: **{statuses['mapped_manufacturer_no_configuration_candidate']:,}**",
        f"- Unmapped manufacturer lineages: **{statuses['unmapped_manufacturer']:,}**",
        "",
        "Candidate counts are discovery metrics, not verified coverage percentages. Exact model names must be obtained from EPA certificates or manufacturer documentation before insertion.",
        "",
        "## Documented Source Corrections",
        "",
        "| Manufacturer | Root family | Raw displacement L | Raw cylinders | Matching correction |",
        "|---|---|---|---|---|",
    ]

    for lineage in sorted(
        overridden_lineages,
        key=lambda item: (item["manufacturer"], item["lineage_root_family"]),
    ):
        override = lineage["matching_override"]
        correction_parts = []
        if "displacements_l" in override:
            correction_parts.append(
                "displacement "
                + ", ".join(str(value) for value in override["displacements_l"])
                + " L"
            )
        if "cylinders" in override:
            correction_parts.append(
                "cylinders "
                + ", ".join(str(value) for value in override["cylinders"])
            )
        lines.append(
            f"| {lineage['manufacturer']} | {lineage['lineage_root_family']} | "
            f"{', '.join(str(value) for value in lineage['displacements_l'])} | "
            f"{', '.join(str(value) for value in lineage['cylinders'])} | "
            f"{'; '.join(correction_parts)} - {override['reason']} |"
        )

    lines.extend(
        [
            "",
            "## Documented Lineage Crosswalks",
            "",
            "| Manufacturer | Root family | Catalog model | Catalog URL | Verification basis |",
            "|---|---|---|---|---|",
        ]
    )
    for lineage in sorted(
        verified_crosswalks,
        key=lambda item: (item["manufacturer"], item["lineage_root_family"]),
    ):
        crosswalk = lineage["verified_lineage_crosswalk"]
        candidate = next(
            (
                item
                for item in lineage["database_candidates"]
                if item["slug"] == crosswalk["slug"]
            ),
            None,
        )
        lines.append(
            f"| {lineage['manufacturer']} | {lineage['lineage_root_family']} | "
            f"{candidate['brand']} {candidate['model']} | "
            f"`/engines/{candidate['slug']}` | {crosswalk['reason']} |"
        )

    lines.extend(
        [
            "",
            "## Manufacturer Baseline",
        "",
        "| EPA manufacturer | Database brand | Annual families | Carryover lineages | Stationary lineages | Stationary strong candidates | Stationary other candidates | Researched exceptions | Stationary no candidate | Stationary unmapped |",
        "|---|---|---:|---:|---:|---:|---:|---:|---:|---:|",
        ]
    )

    for manufacturer, counts in sorted(
        manufacturers.items(),
        key=lambda item: (-item[1]["lineages"], item[0]),
    ):
        brands = ", ".join(sorted(MANUFACTURER_BRANDS.get(manufacturer, set())))
        lines.append(
            f"| {manufacturer} | {brands or 'Unmapped'} | "
            f"{counts['annual_families']} | {counts['lineages']} | "
            f"{counts['stationary_lineages']} | "
            f"{counts['stationary_strong_configuration_candidate']} | "
            f"{counts['stationary_configuration_candidate']} | "
            f"{counts['stationary_researched_configuration_exception']} | "
            f"{counts['stationary_mapped_manufacturer_no_configuration_candidate']} | "
            f"{counts['stationary_unmapped_manufacturer']} |"
        )

    lines.extend(
        [
            "",
            "## Recent Discovery Queue",
            "",
            "These 2024+ stationary lineages have no same-brand displacement, cylinder and aspiration-compatible candidate or use an unmapped certification manufacturer. Mobile-only certification lineages are excluded.",
            "",
            "| Latest year | Manufacturer | Root family | Stationary test | Displacement L | Cylinders | Max power kW | Status |",
            "|---:|---|---|---|---|---|---|---|",
        ]
    )
    for lineage in sorted(
        recent_stationary_gaps,
        key=lambda item: (
            -max(item["model_years"]),
            item["manufacturer"],
            item["lineage_root_family"],
        ),
    )[:200]:
        lines.append(
            f"| {max(lineage['model_years'])} | {lineage['manufacturer']} | "
            f"{lineage['lineage_root_family']} | "
            f"{'Yes' if lineage['has_stationary_test'] else 'No'} | "
            f"{', '.join(str(value) for value in lineage['displacements_l'])} | "
            f"{', '.join(str(value) for value in lineage['cylinders'])} | "
            f"{', '.join(str(value) for value in lineage['max_powers_kw'])} | "
            f"{lineage['coverage_status']} |"
        )
    if not recent_stationary_gaps:
        lines.append("| - | None | None | - | - | - | - | - |")

    lines.extend(
        [
            "",
            "## Remaining Stationary Research Exceptions",
            "",
            f"These {len(researched_stationary_exceptions)} researched stationary lineages have been investigated and deliberately quarantined. They are not claimed as catalog matches because available evidence is insufficient or conflicts with the nearest model. Any remaining unresearched gaps also appear here.",
            "",
            "| Latest year | Manufacturer | Root family | Displacement L | Cylinders | Max power kW | EPA test numbers | Status | Research finding |",
            "|---:|---|---|---|---|---|---|---|---|",
        ]
    )
    for lineage in sorted(
        remaining_stationary_gaps,
        key=lambda item: (
            -max(item["model_years"]),
            item["manufacturer"],
            item["lineage_root_family"],
        ),
    ):
        lines.append(
            f"| {max(lineage['model_years'])} | {lineage['manufacturer']} | "
            f"{lineage['lineage_root_family']} | "
            f"{', '.join(str(value) for value in lineage['displacements_l'])} | "
            f"{', '.join(str(value) for value in lineage['cylinders'])} | "
            f"{', '.join(str(value) for value in lineage['max_powers_kw'])} | "
            f"{', '.join(lineage['cert_test_numbers'])} | "
            f"{lineage['coverage_status']} | "
            f"{RESEARCH_EXCEPTION_NOTES.get((lineage['manufacturer'], lineage['lineage_root_family']), 'Exact commercial model is not publicly identified.')} |"
        )
    if not remaining_stationary_gaps:
        lines.append("| - | None | None | - | - | - | - | - | - |")

    lines.extend(
        [
            "",
            "## Recent Mobile-Only Exclusions",
            "",
            "These 2024+ lineages lack a same-brand catalog configuration candidate, but the EPA workbook classifies them only under mobile, marine, or small-SI equipment test modes. They are intentionally excluded from generator-engine coverage until a separate manufacturer source documents a stationary or generator application.",
            "",
            "| Latest year | Manufacturer | Root family | Displacement L | Cylinders | Max power kW | EPA application/test mode | Status |",
            "|---:|---|---|---|---|---|---|---|",
        ]
    )
    for lineage in sorted(
        recent_mobile_only_gaps,
        key=lambda item: (
            -max(item["model_years"]),
            item["manufacturer"],
            item["lineage_root_family"],
        ),
    ):
        lines.append(
            f"| {max(lineage['model_years'])} | {lineage['manufacturer']} | "
            f"{lineage['lineage_root_family']} | "
            f"{', '.join(str(value) for value in lineage['displacements_l'])} | "
            f"{', '.join(str(value) for value in lineage['cylinders'])} | "
            f"{', '.join(str(value) for value in lineage['max_powers_kw'])} | "
            f"{', '.join(lineage['mobile_stationary_tests'])} | "
            f"{lineage['coverage_status']} |"
        )
    if not recent_mobile_only_gaps:
        lines.append("| - | None | None | - | - | - | - | - |")

    lines.extend(
        [
            "",
            "## Next Step",
            "",
            "Resolve the remaining stationary research exceptions by opening the corresponding EPA certificate or manufacturer certification listing. Add or crosswalk a model only when the exact commercial identity, fuel, displacement, power node and stationary/generator application agree.",
            "",
        ]
    )
    return "\n".join(lines)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--epa-xlsx", required=True, type=Path)
    parser.add_argument("--engines-json", required=True, type=Path)
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("reports/epa-spark-ignited"),
    )
    args = parser.parse_args()

    annual_families, source_rows = read_workbook(args.epa_xlsx)
    lineages = aggregate_lineages(annual_families)
    catalog_bytes = args.engines_json.read_bytes()
    engines = json.loads(catalog_bytes)
    results = classify_lineages(lineages, engines)
    source_sha256 = hashlib.sha256(args.epa_xlsx.read_bytes()).hexdigest()
    catalog_sha256 = hashlib.sha256(catalog_bytes).hexdigest()

    args.output_dir.mkdir(parents=True, exist_ok=True)
    json_path = args.output_dir / "epa-si-1800rpm-lineage-match.json"
    report_path = args.output_dir / "epa-si-1800rpm-summary.md"
    json_path.write_text(json.dumps(results, indent=2, ensure_ascii=False) + "\n")
    report_path.write_text(
        markdown_report(
            annual_families,
            results,
            source_rows,
            args.epa_xlsx,
            source_sha256,
            len(engines),
            catalog_sha256,
        )
    )
    print(f"Wrote {report_path}")
    print(f"Wrote {json_path}")


if __name__ == "__main__":
    main()
